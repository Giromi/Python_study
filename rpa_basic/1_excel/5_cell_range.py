from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

#1줄씩 데이터 넣기
ws.append((["번호", "영어", "수학"]))
for i in range(1, 11):		# 10개 데이터 넣기
	ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]			# 영어 column만 가져오기
# print(col_B)
for cell in col_B:
	print(cell.value)

col_range = ws["B:C"]	# 영어, 수학 column 함께 가지고 오기
for cols in col_range:
	for cell in cols:
		print(cell.value)

print("-------------------------------------------")

row_title = ws[1] # 1번 째 row 만 가져오기
for cell in row_title:
	print(cell.value)

print("-------------------------------------------")

row_range = ws[2:6]			# 6을 포함해서 가져옴
							# 1번 재 줄인 title을 제외하고
for rows in row_range:
	for cell in rows:
		print(cell.value, end=" ")
	print()

print("-------------------------------------------")

row_range = ws[2:ws.max_row]	# 2번 째 줄부터 마지막 줄까지
for rows in row_range:
	for cell in rows:
		print(cell.value, end=" ")
		print(cell.coordinate, end=" ")
	print()
print("-------------------------------------------")

row_range = ws[2:ws.max_row]	# 2번 째 줄부터 마지막 줄까지
for rows in row_range:
	for cell in rows:
		xy = coordinate_from_string(cell.coordinate)
		#  print(xy, end=" ")		# turple
		print(xy[0], end=" ")
		print(xy[1], end=" ")
	print()

print("-------------------------------------------")

#  print(ws.rows)
# whole columns
print(tuple(ws.columns))
print(tuple(ws.rows))

for row in tuple(ws.rows):
	print(row[2].value)

for column in tuple(ws.columns):
	print(column[0].value)

for row in tuple(ws.columns):		# 전체 row
	print(row[0].value)

for column in tuple(ws.columns):	# 전체 column
	print(column[0].value)

for row in ws.iter_row(min_row=1, max=row=5)	 # 전체 row
wb.save("sample.xlsx")

