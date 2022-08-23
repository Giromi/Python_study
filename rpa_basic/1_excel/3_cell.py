from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "Giromi Sheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])					# A1 print cell's information
print(ws["A1"].value)			# A1 print cell's value
print(ws["A10"].value)			# if no value then print 'None'

# row = 1, 2, 3, ...
# col = A(1), B(2), c(3) ...
print(ws.cell(column=1, row=1).value)		# ws["A1"].value
print(ws.cell(column=2, row=1).value)		# ws["B1"].value

c = ws.cell(column=3, row=1, value=10)		# ws["C1"].value = 10
print(c.value)								# ws["C1"]
# 좀 더 반복 넣기 수월함

# 57:00부터 다시 시청

index = 1
for x in range(1, 11):
	for y in range (1, 11):
		#  ws.cell(row=x, column=y, value=randint(0, 100))	#0 ~ 100 number
		ws.cell(row=x, column=y, value=index)
		index += 1



wb.save("sample.xlsx")
