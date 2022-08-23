from openpyxl import Workbook

wb = Workbook()

ws = wb.create_sheet()	# 새로운 Sheet 생성
ws.title = "Giromi Sheet"	# Sheet 이름 변경
ws.sheet_properties.tabcolor = "ff66ff" #RGB 형태로 값을 넣어주면 색상 변경

#sheet, MySheet, YOU
ws1 = wb.create_sheet("Giromi sheet1")	#주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)	#2번 째 idx 에 sheet 생성

new_ws = wb["Giromi Sheet"]
print(wb. sheetnames)					# 모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
taget = wb.copy_worksheet(new_ws)
taget.title = "Copied sheet"


wb.save("sample.xlsx")

